import os
import json
import pandas as pd
from openpyxl import load_workbook

def sanitize_string(s):
    if s is None:
        return ''
    return ''.join(c for c in s if ord(c) in range(32, 127))

base_folder = r'path\to\your\messages\folder'

total_messages = 0
for subfolder in os.listdir(base_folder):
    if subfolder.startswith('c'):
        messages_file = os.path.join(base_folder, subfolder, 'messages.json')
        if os.path.isfile(messages_file):
            with open(messages_file, 'r', encoding='utf-8') as f:
                messages_data = json.load(f)
                total_messages += len(messages_data)

print(f"Total messages to process: {total_messages}")

data = []
processed_messages = 0

for subfolder in os.listdir(base_folder):
    if subfolder.startswith('c'):
        channel_folder = os.path.join(base_folder, subfolder)
        channel_file = os.path.join(channel_folder, 'channel.json')
        messages_file = os.path.join(channel_folder, 'messages.json')

        with open(channel_file, 'r', encoding='utf-8') as f:
            channel_data = json.load(f)

        if 'guild' in channel_data:
            server_name = channel_data['guild']['name']
            channel_name = channel_data['name']
        elif 'recipients' in channel_data:
            server_name = 'Direct Messages'
            channel_name = ', '.join(channel_data['recipients'])
        else:
            server_name = 'Unknown or Deleted'
            channel_name = 'Unknown or Deleted'

        with open(messages_file, 'r', encoding='utf-8') as f:
            messages_data = json.load(f)

        for message in messages_data:
            data.append({
                'Server': sanitize_string(server_name),
                'Channel': sanitize_string(channel_name),
                'Message ID': message.get('ID'),
                'Timestamp': message.get('Timestamp'),
                'Contents': sanitize_string(message.get('Contents')),
                'Attachments': sanitize_string(message.get('Attachments'))
            })
            processed_messages += 1
            print(f"Processed {processed_messages}/{total_messages} messages")

df = pd.DataFrame(data)
excel_path = 'discord_messages.xlsx'
df.to_excel(excel_path, index=False)

wb = load_workbook(excel_path)
ws = wb.active

for cell in ws['C']:
    if cell.row == 1:
        continue
    cell.number_format = '0'

wb.save(excel_path)